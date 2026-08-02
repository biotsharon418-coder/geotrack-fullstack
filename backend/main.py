"""
main.py — GeoTrack API v3
Features: 2FA, OTP email verification, audit logs, session timeout,
          account lockout, password policy, charts data, export endpoints.
"""
import os, secrets, random, string
from datetime import datetime, timedelta
from typing import List, Optional
from io import BytesIO

import pyotp
from fastapi import FastAPI, Depends, HTTPException, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from sqlalchemy import and_, extract
from datetime import datetime, timedelta

from apscheduler.schedulers.background import BackgroundScheduler

import models, schemas
from database import engine, get_db, Base, SessionLocal
from auth import hash_password, verify_password, create_access_token, require_student, require_osas_admin
import email_utils
from models import (
    User,
    BoardingHouse,
    StatusUpdate,
    Review,
    Concern,
    PasswordResetToken,
    AuditLog,
    StudentCompliance,
    StudentFlag,
    Notification,
)

Base.metadata.create_all(bind=engine)

app = FastAPI(title="GeoTrack API", version="3.0.0")

_origins_env = os.getenv("ALLOWED_ORIGINS", "")
# Browsers reject "Access-Control-Allow-Origin: *" together with credentialed
# requests (the frontend sends a Bearer token / cookies), so "*" can never be
# used here. Fall back to the known deployed frontend domains instead of a
# wildcard when ALLOWED_ORIGINS isn't set in the environment.
ALLOWED_ORIGINS = [o.strip() for o in _origins_env.split(",") if o.strip()] or [
    "https://geotrack-osas.vercel.app",
    "https://geotrack-lspu.vercel.app",
    "http://localhost:5173",
]
app.add_middleware(CORSMiddleware, allow_origins=ALLOWED_ORIGINS,
                   allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

def _scheduled_compliance_sweep():
    """Runs on its own DB session since it's not triggered by a request."""
    db = SessionLocal()
    try:
        run_compliance_automation(db, actor=None)
    finally:
        db.close()

scheduler = BackgroundScheduler(timezone="UTC")
scheduler.add_job(_scheduled_compliance_sweep, "cron", hour=0, minute=10, id="compliance_daily_sweep")

@app.on_event("startup")
def _start_scheduler():
    if not scheduler.running:
        scheduler.start()

@app.on_event("shutdown")
def _stop_scheduler():
    if scheduler.running:
        scheduler.shutdown(wait=False)

MAX_FAILED = 5
LOCKOUT_MINS = 5
ARCHIVE_YEARS = 3
DELETE_YEARS = 5

COMPLIANCE_DEADLINE_DAY = 25


def check_monthly_compliance(db: Session):

    now = datetime.utcnow()

    month = now.strftime("%B")
    year = now.year

    students = (
        db.query(models.User)
        .filter(models.User.role == "student")
        .all()
    )

    for student in students:

        compliance = (
            db.query(models.StudentCompliance)
            .filter(
                models.StudentCompliance.student_id == student.id,
                models.StudentCompliance.month == month,
                models.StudentCompliance.year == year,
            )
            .first()
        )

        if not compliance:

            compliance = models.StudentCompliance(
                student_id=student.id,
                month=month,
                year=year,
                submission_status="Pending",
                deadline=datetime(year, now.month, COMPLIANCE_DEADLINE_DAY),
                remarks="Waiting for submission",
            )

            db.add(compliance)

        if (
            compliance.submission_status == "Pending"
            and now.day > COMPLIANCE_DEADLINE_DAY
        ):
            compliance.submission_status = "Missed"
            compliance.remarks = "Submission deadline missed"

            flag = (
                db.query(models.StudentFlag)
                .filter(models.StudentFlag.student_id == student.id)
                .first()
            )

            if not flag:
                flag = models.StudentFlag(
                    student_id=student.id,
                    missed_count=0,
                    compliance_status="Good",
                )
                db.add(flag)

            flag.missed_count += 1
            was_flagged = flag.is_flagged

            if flag.missed_count >= 3:
                flag.is_flagged = True
                flag.compliance_status = "Flagged"
                if not was_flagged:
                    db.commit()
                    email_utils.notify_flagged(student.email, student.full_name, flag.missed_count)
                    notify(db, student.id, "flagging", "You've been flagged for compliance",
                           f"You missed {flag.missed_count} monthly submission(s) and have been flagged by OSAS. "
                           "Please submit your status update as soon as possible.")

    db.commit()

# ─── Helpers ─────────────────────────────────────────────────────────────────
def log_action(db: Session, actor: Optional[models.User], action: str,
               resource_type: str, resource_id: Optional[int] = None,
               resource_label: Optional[str] = None, detail: Optional[str] = None):
    db.add(models.AuditLog(
        actor_id=actor.id if actor else None,
        actor_name=actor.full_name if actor else "System",
        actor_role=actor.role if actor else "system",
        action=action, resource_type=resource_type,
        resource_id=resource_id, resource_label=resource_label, detail=detail,
    ))
    db.commit()


def notify(db: Session, user_id: Optional[int], category: str, title: str, message: str):
    """Create an in-app notification for one user. Safe no-op if user_id is None."""
    if not user_id:
        return
    db.add(models.Notification(user_id=user_id, category=category, title=title, message=message))
    db.commit()


def notify_all(db: Session, user_ids, category: str, title: str, message: str):
    for uid in user_ids:
        db.add(models.Notification(user_id=uid, category=category, title=title, message=message))
    db.commit()


def make_otp() -> str:
    return "".join(random.choices(string.digits, k=6))


def lifecycle_sweep(db: Session):
    now = datetime.utcnow()
    for s in db.query(models.User).filter(models.User.role == "student").all():
        if s.archived_at is None:
            last = (db.query(models.StatusUpdate)
                    .filter(models.StatusUpdate.student_id == s.id,
                            models.StatusUpdate.status_type.in_(["same", "transferred"]))
                    .order_by(models.StatusUpdate.created_at.desc()).first())
            if (last.created_at if last else s.created_at) < now - timedelta(days=365*ARCHIVE_YEARS):
                s.archived_at = now
        elif s.archived_at < now - timedelta(days=365*DELETE_YEARS):
            db.delete(s)
    db.commit()


# ─── AUTH ─────────────────────────────────────────────────────────────────────
@app.post("/api/auth/register/student", response_model=schemas.TokenResponse)
def register_student(payload: schemas.RegisterStudentRequest, db: Session = Depends(get_db)):
    if db.query(models.User).filter(models.User.email == payload.email).first():
        raise HTTPException(400, "Email already registered.")
    otp = make_otp()
    user = models.User(
        full_name=payload.full_name, email=payload.email,
        hashed_password=hash_password(payload.password),
        role="student", course_section=payload.course_section, gender=payload.gender,
        is_email_verified=False,
        email_otp=otp, email_otp_expires=datetime.utcnow() + timedelta(minutes=15),
    )
    db.add(user); db.commit(); db.refresh(user)
    email_utils.notify_email_otp(user.email, user.full_name, otp)

    if payload.boarding_house_name and payload.boarding_house_barangay:
        house = db.query(models.BoardingHouse).filter(
            models.BoardingHouse.name == payload.boarding_house_name).first()
        if not house:
            house = models.BoardingHouse(
                name=payload.boarding_house_name, barangay=payload.boarding_house_barangay,
                latitude=payload.boarding_house_latitude, longitude=payload.boarding_house_longitude,
                is_verified=False, submitted_by=f"Student — {user.full_name}",
                submitted_by_id=user.id)
            db.add(house); db.commit(); db.refresh(house)
        db.add(models.StatusUpdate(student_id=user.id, boarding_house_id=house.id,
                                   status_type="same", month_label=datetime.utcnow().strftime("%B %Y")))
        db.commit()

        current = datetime.utcnow()

        db.add(
            models.StudentCompliance(
                student_id=user.id,
                month=current.strftime("%B"),
                year=current.year,
                submission_status="Submitted",
                submitted_at=current,
                deadline=datetime(
                    current.year,
                    current.month,
                    COMPLIANCE_DEADLINE_DAY
                ),
                remarks="Initial registration"
            )
        )

        db.commit()

    log_action(db, user, "create", "user", user.id, user.full_name, "Student registered")
    token = create_access_token({"sub": str(user.id), "role": user.role})
    # requires_otp=True tells frontend to show the OTP verification screen
    return schemas.TokenResponse(access_token=token, role=user.role,
                                 full_name=user.full_name, requires_otp=True,
                                 pending_token=otp)   # demo: return OTP so UI can prefill it


@app.post("/api/auth/verify-otp")
def verify_email_otp(payload: schemas.VerifyOTPRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == payload.email).first()
    if not user or user.email_otp != payload.otp:
        raise HTTPException(400, "Invalid OTP code.")
    if user.email_otp_expires and user.email_otp_expires < datetime.utcnow():
        raise HTTPException(400, "OTP has expired. Please request a new one.")
    user.is_email_verified = True
    user.email_otp = None; user.email_otp_expires = None
    db.commit()
    log_action(db, user, "verify", "email", user.id, user.email, "Email OTP verified")
    return {"message": "Email verified successfully."}


@app.post("/api/auth/resend-otp")
def resend_otp(email: str, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == email).first()
    if not user: raise HTTPException(404, "User not found.")
    otp = make_otp()
    user.email_otp = otp; user.email_otp_expires = datetime.utcnow() + timedelta(minutes=15)
    db.commit()
    sent = email_utils.notify_email_otp(user.email, user.full_name, otp)
    resp = {"message": "New OTP sent." if sent else "New OTP generated (email not configured)."}
    if not sent:
        resp["demo_otp"] = otp
    return resp


@app.post("/api/auth/register/osas", response_model=schemas.TokenResponse)
def register_osas(payload: schemas.RegisterOsasRequest, db: Session = Depends(get_db)):
    if db.query(models.User).filter(models.User.email == payload.email).first():
        raise HTTPException(400, "Email already registered.")
    user = models.User(full_name=payload.full_name, email=payload.email,
                       hashed_password=hash_password(payload.password),
                       role="osas_admin", position=payload.position,
                       is_email_verified=True)
    db.add(user); db.commit(); db.refresh(user)
    log_action(db, user, "create", "user", user.id, user.full_name, "OSAS admin registered")
    token = create_access_token({"sub": str(user.id), "role": user.role})
    return schemas.TokenResponse(access_token=token, role=user.role, full_name=user.full_name)


@app.post("/api/auth/token", response_model=schemas.TokenResponse)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == form_data.username).first()
    generic = "Incorrect email or password."
    if not user: raise HTTPException(401, generic)
    if user.locked_until and user.locked_until > datetime.utcnow():
        mins = int((user.locked_until - datetime.utcnow()).total_seconds() // 60) + 1
        raise HTTPException(429, f"Account locked for {mins} more minute(s).")
    if not verify_password(form_data.password, user.hashed_password):
        user.failed_login_attempts = (user.failed_login_attempts or 0) + 1
        if user.failed_login_attempts >= MAX_FAILED:
            user.locked_until = datetime.utcnow() + timedelta(minutes=LOCKOUT_MINS)
            user.failed_login_attempts = 0
            db.commit()
            log_action(db, user, "lockout", "user", user.id, user.email, f"Locked after {MAX_FAILED} failed attempts")
            raise HTTPException(429, f"Too many failed attempts. Account locked for {LOCKOUT_MINS} minutes.")
        db.commit()
        raise HTTPException(401, generic)

    user.failed_login_attempts = 0; user.locked_until = None; db.commit()

    # 2FA check
    if user.two_fa_enabled:
        pending = secrets.token_urlsafe(24)
        user.pending_2fa_token = pending
        user.pending_2fa_expires = datetime.utcnow() + timedelta(minutes=5)
        db.commit()
        return schemas.TokenResponse(access_token="", role=user.role,
                                     full_name=user.full_name, requires_2fa=True,
                                     pending_token=pending)

    log_action(db, user, "login", "user", user.id, user.full_name)
    token = create_access_token({"sub": str(user.id), "role": user.role})
    return schemas.TokenResponse(access_token=token, role=user.role, full_name=user.full_name)


@app.post("/api/auth/2fa/verify", response_model=schemas.TokenResponse)
def verify_2fa(payload: schemas.TwoFAVerifyRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.pending_2fa_token == payload.pending_token).first()
    if not user: raise HTTPException(400, "Invalid or expired 2FA session.")
    if user.pending_2fa_expires and user.pending_2fa_expires < datetime.utcnow():
        raise HTTPException(400, "2FA session expired. Please sign in again.")
    totp = pyotp.TOTP(user.totp_secret)
    if not totp.verify(payload.totp_code, valid_window=1):
        raise HTTPException(400, "Invalid 2FA code.")
    user.pending_2fa_token = None; user.pending_2fa_expires = None; db.commit()
    log_action(db, user, "login", "user", user.id, user.full_name, "2FA verified")
    token = create_access_token({"sub": str(user.id), "role": user.role})
    return schemas.TokenResponse(access_token=token, role=user.role, full_name=user.full_name)


@app.post("/api/auth/2fa/setup", response_model=schemas.TwoFASetupResponse)
def setup_2fa(db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    secret = pyotp.random_base32()
    user.totp_secret = secret; db.commit()
    totp = pyotp.TOTP(secret)
    uri = totp.provisioning_uri(name=user.email, issuer_name="GeoTrack LSPU")
    return schemas.TwoFASetupResponse(secret=secret, qr_uri=uri)


@app.post("/api/auth/2fa/enable")
def enable_2fa(code: str, db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    if not user.totp_secret: raise HTTPException(400, "Run /api/auth/2fa/setup first.")
    totp = pyotp.TOTP(user.totp_secret)
    if not totp.verify(code, valid_window=1): raise HTTPException(400, "Invalid TOTP code.")
    user.two_fa_enabled = True; db.commit()
    log_action(db, user, "update", "user", user.id, user.email, "2FA enabled")
    return {"message": "2FA enabled."}


@app.post("/api/auth/2fa/disable")
def disable_2fa(db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    user.two_fa_enabled = False; user.totp_secret = None; db.commit()
    log_action(db, user, "update", "user", user.id, user.email, "2FA disabled")
    return {"message": "2FA disabled."}


@app.post("/api/auth/forgot-password")
def forgot_password(payload: schemas.ForgotPasswordRequest, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == payload.email).first()
    if not user:
        return {"message": "If that email is registered, a reset link has been sent.", "demo_token": None}
    db.query(models.PasswordResetToken).filter(
        models.PasswordResetToken.user_id == user.id,
        models.PasswordResetToken.used == False).delete()
    db.commit()
    token = secrets.token_urlsafe(32)
    db.add(models.PasswordResetToken(user_id=user.id, token=token,
                                     expires_at=datetime.utcnow() + timedelta(hours=1)))
    db.commit()
    return {"message": "Reset token generated.", "demo_token": token,
            "demo_note": "In production this is emailed. Copy and paste into the reset form."}


@app.post("/api/auth/reset-password")
def reset_password(payload: schemas.ResetPasswordRequest, db: Session = Depends(get_db)):
    record = db.query(models.PasswordResetToken).filter(
        models.PasswordResetToken.token == payload.token,
        models.PasswordResetToken.used == False).first()
    if not record: raise HTTPException(400, "Invalid or already-used reset token.")
    if record.expires_at < datetime.utcnow(): raise HTTPException(400, "Token expired.")
    record.user.hashed_password = hash_password(payload.new_password)
    record.user.failed_login_attempts = 0; record.user.locked_until = None
    record.used = True; db.commit()
    log_action(db, record.user, "update", "user", record.user.id, record.user.email, "Password reset via token")
    return {"message": "Password updated. You can now sign in."}


# ─── STUDENT endpoints ────────────────────────────────────────────────────────
@app.get("/api/student/me", response_model=schemas.MyProfileOut)
def my_profile(user: models.User = Depends(require_student)):
    return user

@app.put("/api/student/me", response_model=schemas.MyProfileOut)
def update_profile(payload: schemas.MyProfileUpdate, db: Session = Depends(get_db),
                   user: models.User = Depends(require_student)):
    old_name = user.full_name
    for f, v in payload.dict(exclude_unset=True).items(): setattr(user, f, v)
    db.commit(); db.refresh(user)
    log_action(db, user, "update", "user", user.id, user.full_name,
               f"Profile updated (was: {old_name})" if old_name != user.full_name else "Profile updated")
    return user

@app.get("/api/student/boarding-houses", response_model=List[schemas.BoardingHouseOut])
def student_list_bh(db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    return db.query(models.BoardingHouse).all()

@app.get("/api/student/boarding-houses/{hid}/reviews", response_model=List[schemas.ReviewOut])
def student_get_reviews(hid: int, db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    if not db.query(models.BoardingHouse).get(hid): raise HTTPException(404, "Not found")
    return db.query(models.Review).filter(models.Review.boarding_house_id == hid).all()

@app.post("/api/student/boarding-houses/{hid}/reviews", response_model=schemas.ReviewOut)
def student_post_review(hid: int, payload: schemas.ReviewCreate, db: Session = Depends(get_db),
                        user: models.User = Depends(require_student)):
    h = db.query(models.BoardingHouse).get(hid)
    if not h: raise HTTPException(404, "Not found")
    if not (1 <= payload.rating <= 5): raise HTTPException(400, "Rating must be 1–5")
    r = models.Review(boarding_house_id=hid, author_id=user.id, rating=payload.rating, text=payload.text)
    db.add(r); db.commit(); db.refresh(r)
    log_action(db, user, "create", "review", r.id, h.name, f"Rating: {payload.rating}/5")
    return r

@app.get("/api/student/my-reviews", response_model=List[schemas.ReviewOut])
def my_reviews(db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    return db.query(models.Review).filter(models.Review.author_id == user.id).all()

@app.put("/api/student/reviews/{rid}", response_model=schemas.ReviewOut)
def update_review(rid: int, payload: schemas.ReviewUpdate, db: Session = Depends(get_db),
                  user: models.User = Depends(require_student)):
    r = db.query(models.Review).get(rid)
    if not r: raise HTTPException(404, "Not found")
    if r.author_id != user.id: raise HTTPException(403, "Not your review")
    if payload.rating is not None:
        if not (1 <= payload.rating <= 5): raise HTTPException(400, "Rating 1–5")
        r.rating = payload.rating
    if payload.text is not None: r.text = payload.text
    db.commit(); db.refresh(r)
    log_action(db, user, "update", "review", r.id, None, "Review updated")
    return r

@app.delete("/api/student/reviews/{rid}")
def delete_review(rid: int, db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    r = db.query(models.Review).get(rid)
    if not r: raise HTTPException(404, "Not found")
    if r.author_id != user.id: raise HTTPException(403, "Not your review")
    db.delete(r); db.commit()
    log_action(db, user, "delete", "review", rid, None, "Review deleted")
    return {"message": "Deleted"}

@app.post("/api/student/status-updates", response_model=schemas.StatusUpdateOut)
def submit_status(
    payload: schemas.StatusUpdateCreate,
    db: Session = Depends(get_db),
    user: models.User = Depends(require_student)
):
    if payload.status_type == "transferred" and not payload.new_boarding_house_name:
        raise HTTPException(
            status_code=400,
            detail="new_boarding_house_name required when transferred"
        )

    current_month = datetime.utcnow().strftime("%B")
    current_year = datetime.utcnow().year

    existing = (
        db.query(models.StudentCompliance)
        .filter(
            models.StudentCompliance.student_id == user.id,
            models.StudentCompliance.month == current_month,
            models.StudentCompliance.year == current_year,
        )
        .first()
    )

    if existing:

        if existing.submission_status == "Submitted":
            raise HTTPException(
                status_code=400,
                detail="You have already submitted your boarding status for this month."
            )

        existing.submission_status = "Submitted"
        existing.submitted_at = datetime.utcnow()
        existing.remarks = "Submitted on time"

    else:

        existing = models.StudentCompliance(
            student_id=user.id,
            month=current_month,
            year=current_year,
            submission_status="Submitted",
            submitted_at=datetime.utcnow(),
            deadline=datetime(
                current_year,
                datetime.utcnow().month,
                COMPLIANCE_DEADLINE_DAY,
            ),
            remarks="Submitted on time",
        )

        db.add(existing)

    update = models.StatusUpdate(
        student_id=user.id,
        boarding_house_id=payload.boarding_house_id,
        status_type=payload.status_type,
        new_boarding_house_name=payload.new_boarding_house_name,
        new_barangay=payload.new_barangay,
        note=payload.note,
        month_label=f"{current_month} {current_year}",
    )

    db.add(update)
    db.commit()
    db.refresh(update)

    log_action(
        db,
        user,
        "create",
        "status_update",
        update.id,
        update.month_label,
        f"Status: {payload.status_type}"
    )

    return update

@app.get("/api/student/status-updates", response_model=List[schemas.StatusUpdateOut])
def my_status_updates(db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    return (db.query(models.StatusUpdate).filter(models.StatusUpdate.student_id == user.id)
            .order_by(models.StatusUpdate.created_at.desc()).all())

@app.put("/api/student/status-updates/{uid}", response_model=schemas.StatusUpdateOut)
def edit_status(uid: int, payload: schemas.StatusUpdateEdit, db: Session = Depends(get_db),
                user: models.User = Depends(require_student)):
    u = db.get(models.StatusUpdate, uid)
    if not u: raise HTTPException(404, "Not found")
    if u.student_id != user.id: raise HTTPException(403, "Not yours")
    for f, v in payload.dict(exclude_unset=True).items(): setattr(u, f, v)
    db.commit(); db.refresh(u)
    log_action(db, user, "update", "status_update", uid, u.month_label, "Status update edited")
    return u

@app.delete("/api/student/status-updates/{uid}")
def delete_status(uid: int, db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    u = db.query(models.StatusUpdate).get(uid)
    if not u: raise HTTPException(404, "Not found")
    if u.student_id != user.id: raise HTTPException(403, "Not yours")
    db.delete(u); db.commit()
    log_action(db, user, "delete", "status_update", uid, None, "Status update deleted")
    return {"message": "Deleted"}

@app.post("/api/student/concerns", response_model=schemas.ConcernOut)
def report_concern(payload: schemas.ConcernCreate, db: Session = Depends(get_db),
                   user: models.User = Depends(require_student)):
    c = models.Concern(student_id=user.id, category=payload.category, details=payload.details)
    db.add(c); db.commit(); db.refresh(c)
    log_action(db, user, "create", "concern", c.id, payload.category, payload.details[:80])
    return c

@app.get("/api/student/my-boarding-house", response_model=Optional[schemas.BoardingHouseOut])
def my_boarding_house(db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    u = (db.query(models.StatusUpdate)
         .filter(models.StatusUpdate.student_id == user.id,
                 models.StatusUpdate.boarding_house_id.isnot(None))
         .order_by(models.StatusUpdate.created_at.desc()).first())
    return u.boarding_house if u else None


# ─── OSAS endpoints ───────────────────────────────────────────────────────────

EMERGENCY_CATEGORIES = ["Medical Emergency", "Safety Threat", "Fire", "Natural Disaster", "Other"]

def _log_timeline(db, case, actor, event, note=None):
    db.add(models.EmergencyTimelineEntry(
        case_id=case.id,
        actor_name=actor.full_name if actor else "System",
        actor_role=actor.role if actor else "system",
        event=event, note=note,
    ))
    db.commit()

def _emergency_out(case):
    return schemas.EmergencyCaseOut(
        id=case.id, student_id=case.student_id,
        student_name=case.student.full_name if case.student else None,
        student_email=case.student.email if case.student else None,
        category=case.category, details=case.details,
        latitude=case.latitude, longitude=case.longitude,
        status=case.status, created_at=case.created_at, resolved_at=case.resolved_at,
        timeline=list(case.timeline),
    )

@app.post("/api/student/sos", response_model=schemas.EmergencyCaseOut)
def trigger_sos(payload: schemas.SOSCreate, db: Session = Depends(get_db),
                user: models.User = Depends(require_student)):
    if payload.category not in EMERGENCY_CATEGORIES:
        raise HTTPException(400, "Invalid emergency category")
    case = models.EmergencyCase(
        student_id=user.id, category=payload.category, details=payload.details,
        latitude=payload.latitude, longitude=payload.longitude, status="Active",
    )
    db.add(case); db.commit(); db.refresh(case)
    _log_timeline(db, case, user, "SOS triggered", payload.details)
    log_action(db, user, "create", "emergency", case.id, payload.category, "SOS triggered")
    osas_admins = db.query(models.User).filter(models.User.role == "osas_admin").all()
    for admin in osas_admins:
        email_utils.notify_emergency_alert(admin.email, payload.category, user.full_name, payload.details)
    notify_all(db, [a.id for a in osas_admins], "sos_alert", f"SOS: {payload.category}",
               f"{user.full_name} triggered an SOS alert ({payload.category}).")
    db.refresh(case)
    return _emergency_out(case)

@app.get("/api/student/sos", response_model=List[schemas.EmergencyCaseOut])
def my_emergencies(db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    cases = (db.query(models.EmergencyCase)
             .filter(models.EmergencyCase.student_id == user.id)
             .order_by(models.EmergencyCase.created_at.desc()).all())
    return [_emergency_out(c) for c in cases]

@app.get("/api/student/sos/{case_id}", response_model=schemas.EmergencyCaseOut)
def my_emergency_detail(case_id: int, db: Session = Depends(get_db),
                        user: models.User = Depends(require_student)):
    case = db.query(models.EmergencyCase).get(case_id)
    if not case or case.student_id != user.id: raise HTTPException(404, "Not found")
    return _emergency_out(case)

@app.patch("/api/student/sos/{case_id}/cancel", response_model=schemas.EmergencyCaseOut)
def cancel_my_emergency(case_id: int, db: Session = Depends(get_db),
                        user: models.User = Depends(require_student)):
    case = db.query(models.EmergencyCase).get(case_id)
    if not case or case.student_id != user.id: raise HTTPException(404, "Not found")
    if case.status in ("Resolved", "Cancelled"): raise HTTPException(400, "Case already closed")
    case.status = "Cancelled"; case.resolved_at = datetime.utcnow(); db.commit()
    _log_timeline(db, case, user, "Cancelled by student")
    log_action(db, user, "update", "emergency", case.id, case.category, "Cancelled by student")
    db.refresh(case)
    return _emergency_out(case)

@app.get("/api/osas/dashboard", response_model=schemas.DashboardStats)
def dashboard(db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    students = db.query(models.User).filter(models.User.role == "student").all()
    updates  = db.query(models.StatusUpdate).all()
    check_monthly_compliance(db)

    # Charts data
    gender_counts: dict = {}
    dept_counts:   dict = {}
    for s in students:
        g = (s.gender or "not_specified").replace("_", " ").title()
        gender_counts[g] = gender_counts.get(g, 0) + 1
        d = s.course_section or "Not specified"
        dept_counts[d] = dept_counts.get(d, 0) + 1

    bar_counts: dict = {}
    status_counts: dict = {}
    for u in updates:
        if u.boarding_house:
            b = u.boarding_house.barangay
            bar_counts[b] = bar_counts.get(b, 0) + 1
        st = {"same":"Same house","transferred":"Transferred","moved_home":"Moved home"}.get(u.status_type, u.status_type)
        status_counts[st] = status_counts.get(st, 0) + 1

    # Recent activities (last 10 audit log entries)
    recent = (db.query(models.AuditLog)
              .order_by(models.AuditLog.created_at.desc()).limit(10).all())
    activities = [{"id":r.id,"actor":r.actor_name,"action":r.action,
                   "resource_type":r.resource_type,"resource_label":r.resource_label,
                   "detail":r.detail,"created_at":r.created_at.isoformat()} for r in recent]

    return schemas.DashboardStats(
        total_students=len(students),
        updates_submitted=len(updates),
        flagged_students=db.query(models.StatusUpdate).filter(models.StatusUpdate.is_flagged==True).count(),
        pending_verifications=db.query(models.BoardingHouse).filter(models.BoardingHouse.is_verified==False).count(),
        active_emergencies=db.query(models.EmergencyCase).filter(models.EmergencyCase.status.in_(["Active","Responding"])).count(),
        by_gender=[{"label":k,"count":v} for k,v in sorted(gender_counts.items())],
        by_department=[{"label":k,"count":v} for k,v in sorted(dept_counts.items())],
        by_barangay=[{"label":k,"count":v} for k,v in sorted(bar_counts.items())],
        by_status=[{"label":k,"count":v} for k,v in sorted(status_counts.items())],
        recent_activities=activities,
    )

@app.get("/api/osas/compliance/dashboard")
def compliance_dashboard(
    db: Session = Depends(get_db),
    user: models.User = Depends(require_osas_admin)
):
    now = datetime.utcnow()

    total_students = db.query(models.User).filter(
        models.User.role == "student"
    ).count()

    submitted = db.query(models.StudentCompliance).filter(
        models.StudentCompliance.month == now.strftime("%B"),
        models.StudentCompliance.year == now.year,
        models.StudentCompliance.submission_status == "Submitted"
    ).count()

    pending = db.query(models.StudentCompliance).filter(
        models.StudentCompliance.month == now.strftime("%B"),
        models.StudentCompliance.year == now.year,
        models.StudentCompliance.submission_status == "Pending"
    ).count()

    missed = db.query(models.StudentCompliance).filter(
        models.StudentCompliance.month == now.strftime("%B"),
        models.StudentCompliance.year == now.year,
        models.StudentCompliance.submission_status == "Missed"
    ).count()

    flagged = db.query(models.StudentFlag).filter(
        models.StudentFlag.is_flagged == True
    ).count()

    recent = (
        db.query(models.StudentCompliance)
        .order_by(models.StudentCompliance.id.desc())
        .limit(10)
        .all()
    )

    return {
        "total_students": total_students,
        "submitted": submitted,
        "pending": pending,
        "missed": missed,
        "flagged": flagged,
        "recent": [
            {
                "student_id": r.student_id,
                "student_name": r.student.full_name,
                "month": r.month,
                "year": r.year,
                "status": r.submission_status,
                "submitted_at": r.submitted_at,
            }
            for r in recent
        ]
    }

@app.get("/api/osas/geo-map", response_model=List[schemas.StudentMapPoint])
def geo_map(db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    points = []
    for s in db.query(models.User).filter(models.User.role == "student").all():
        u = (db.query(models.StatusUpdate)
             .filter(models.StatusUpdate.student_id == s.id,
                     models.StatusUpdate.boarding_house_id.isnot(None))
             .order_by(models.StatusUpdate.created_at.desc()).first())
        if not u or not u.boarding_house: continue
        h = u.boarding_house
        if h.latitude is None: continue
        points.append(schemas.StudentMapPoint(student_name=s.full_name, boarding_house_name=h.name,
                                              barangay=h.barangay, latitude=h.latitude,
                                              longitude=h.longitude, is_flagged=u.is_flagged))
    return points

@app.get("/api/osas/status-updates", response_model=List[schemas.StatusUpdateAdminOut])
def all_status_updates(
    search: Optional[str] = None,
    is_flagged: Optional[bool] = None,
    gender: Optional[str] = None,
    month_label: Optional[str] = None,
    is_verified: Optional[bool] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(require_osas_admin),
):
    q = db.query(models.StatusUpdate)
    if is_flagged is not None: q = q.filter(models.StatusUpdate.is_flagged == is_flagged)
    if month_label: q = q.filter(models.StatusUpdate.month_label == month_label)
    updates = q.order_by(models.StatusUpdate.created_at.desc()).all()
    results = []
    for u in updates:
        if not u.student: continue
        if search and search.lower() not in u.student.full_name.lower() and search.lower() not in u.student.email.lower(): continue
        if gender and (u.student.gender or "").lower() != gender.lower(): continue
        if is_verified is not None and u.boarding_house and u.boarding_house.is_verified != is_verified: continue
        results.append(schemas.StatusUpdateAdminOut(
            id=u.id, status_type=u.status_type,
            new_boarding_house_name=u.new_boarding_house_name, new_barangay=u.new_barangay,
            note=u.note, month_label=u.month_label, is_flagged=u.is_flagged,
            flag_reason=u.flag_reason, created_at=u.created_at,
            student_name=u.student.full_name, student_email=u.student.email,
        ))
    return results

@app.patch("/api/osas/status-updates/{uid}/flag")
def flag_update(
    uid: int,
    reason: str,
    db: Session = Depends(get_db),
    user: models.User = Depends(require_osas_admin),
):
    u = db.get(models.StatusUpdate, uid)

    if not u:
        raise HTTPException(404, "Not found")

    u.is_flagged = True
    u.flag_reason = reason
    db.commit()

    notify(db, u.student_id, "flagging", "Your status update was flagged",
           f"Reason: {reason}")

    log_action(
        db,
        user,
        "flag",
        "status_update",
        uid,
        u.student.full_name if u.student else None,
        f"Reason: {reason}",
    )

    return {"message": "Flagged"}

@app.get("/api/osas/boarding-houses", response_model=List[schemas.BoardingHouseOut])
def list_bh(db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    return db.query(models.BoardingHouse).all()

@app.patch("/api/osas/boarding-houses/{hid}/verify", response_model=schemas.BoardingHouseOut)
def verify_bh(hid: int, db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    h = db.query(models.BoardingHouse).get(hid)
    if not h: raise HTTPException(404, "Not found")
    h.is_verified = True; db.commit(); db.refresh(h)
    log_action(db, user, "verify", "boarding_house", hid, h.name, "Boarding house verified")
    if h.submitter:
        email_utils.notify_boarding_house_approved(h.submitter.email, h.submitter.full_name, h.name)
        notify(db, h.submitter.id, "approval", "Boarding house approved",
               f"Your submission for \"{h.name}\" has been verified by OSAS.")
    return h

@app.post("/api/osas/boarding-houses/{hid}/reject", response_model=schemas.BoardingHouseOut)
def reject_bh(hid: int, payload: schemas.BoardingHouseReject, db: Session = Depends(get_db),
             user: models.User = Depends(require_osas_admin)):
    h = db.query(models.BoardingHouse).get(hid)
    if not h: raise HTTPException(404, "Not found")
    h.is_verified = False; db.commit(); db.refresh(h)
    log_action(db, user, "reject", "boarding_house", hid, h.name, payload.reason or "Boarding house rejected")
    if h.submitter:
        email_utils.notify_boarding_house_rejected(h.submitter.email, h.submitter.full_name, h.name, payload.reason)
        notify(db, h.submitter.id, "rejection", "Boarding house rejected",
               f"Your submission for \"{h.name}\" was rejected. Reason: {payload.reason or 'Not specified'}")
    return h

@app.put("/api/osas/boarding-houses/{hid}", response_model=schemas.BoardingHouseOut)
def update_bh(hid: int, payload: schemas.BoardingHouseUpdate, db: Session = Depends(get_db),
              user: models.User = Depends(require_osas_admin)):
    h = db.query(models.BoardingHouse).get(hid)
    if not h: raise HTTPException(404, "Not found")
    for f, v in payload.dict(exclude_unset=True).items(): setattr(h, f, v)
    db.commit(); db.refresh(h)
    log_action(db, user, "update", "boarding_house", hid, h.name, "Details updated")
    return h

@app.delete("/api/osas/boarding-houses/{hid}")
def delete_bh(hid: int, db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    h = db.query(models.BoardingHouse).get(hid)
    if not h: raise HTTPException(404, "Not found")
    log_action(db, user, "delete", "boarding_house", hid, h.name, "Boarding house deleted")
    db.delete(h); db.commit()
    return {"message": "Deleted"}

@app.get("/api/osas/boarding-houses/{hid}/reviews", response_model=List[schemas.ReviewOut])
def bh_reviews(hid: int, db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    if not db.query(models.BoardingHouse).get(hid): raise HTTPException(404, "Not found")
    return db.query(models.Review).filter(models.Review.boarding_house_id == hid).all()

@app.get("/api/osas/concerns", response_model=List[schemas.ConcernAdminOut])
def all_concerns(db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    return [schemas.ConcernAdminOut(id=c.id, category=c.category, details=c.details,
                                    status=c.status, created_at=c.created_at,
                                    student_name=c.student.full_name, student_email=c.student.email)
            for c in db.query(models.Concern).order_by(models.Concern.created_at.desc()).all()]

@app.patch("/api/osas/concerns/{cid}/status")
def update_concern(cid: int, new_status: str, db: Session = Depends(get_db),
                   user: models.User = Depends(require_osas_admin)):
    if new_status not in ("open","in_progress","resolved"): raise HTTPException(400,"Invalid")
    c = db.query(models.Concern).get(cid)
    if not c: raise HTTPException(404, "Not found")
    c.status = new_status; db.commit()
    if new_status == "resolved" and c.student_id:
        notify(db, c.student_id, "resolution", "Your concern was resolved",
               f"Your reported concern ({c.category}) has been marked resolved by OSAS.")
    log_action(db, user, "update", "concern", cid,
               c.student.full_name if c.student else None, f"Status → {new_status}")
    return {"message": "Updated"}

@app.delete("/api/osas/concerns/{cid}")
def delete_concern(cid: int, db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    c = db.query(models.Concern).get(cid)
    if not c: raise HTTPException(404, "Not found")
    log_action(db, user, "delete", "concern", cid,
               c.student.full_name if c.student else None, "Concern deleted")
    db.delete(c); db.commit()
    return {"message": "Deleted"}

@app.get("/api/osas/emergencies", response_model=List[schemas.EmergencyCaseOut])
def list_emergencies(status: Optional[str] = None, db: Session = Depends(get_db),
                     user: models.User = Depends(require_osas_admin)):
    q = db.query(models.EmergencyCase)
    if status: q = q.filter(models.EmergencyCase.status == status)
    cases = q.order_by(models.EmergencyCase.created_at.desc()).all()
    return [_emergency_out(c) for c in cases]

@app.get("/api/osas/emergencies/{case_id}", response_model=schemas.EmergencyCaseOut)
def emergency_detail(case_id: int, db: Session = Depends(get_db),
                     user: models.User = Depends(require_osas_admin)):
    case = db.query(models.EmergencyCase).get(case_id)
    if not case: raise HTTPException(404, "Not found")
    return _emergency_out(case)

@app.patch("/api/osas/emergencies/{case_id}/status", response_model=schemas.EmergencyCaseOut)
def update_emergency_status(case_id: int, payload: schemas.EmergencyStatusUpdate,
                            db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    if payload.status not in ("Active","Responding","Resolved","Cancelled"):
        raise HTTPException(400, "Invalid status")
    case = db.query(models.EmergencyCase).get(case_id)
    if not case: raise HTTPException(404, "Not found")
    case.status = payload.status
    if payload.status in ("Resolved","Cancelled"): case.resolved_at = datetime.utcnow()
    db.commit()
    _log_timeline(db, case, user, f"Status: {payload.status}", payload.note)
    log_action(db, user, "update", "emergency", case.id,
              case.student.full_name if case.student else None, f"Status \u2192 {payload.status}")
    db.refresh(case)
    return _emergency_out(case)

@app.post("/api/osas/emergencies/{case_id}/notes", response_model=schemas.EmergencyCaseOut)
def add_emergency_note(case_id: int, payload: schemas.EmergencyNoteCreate,
                       db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    case = db.query(models.EmergencyCase).get(case_id)
    if not case: raise HTTPException(404, "Not found")
    _log_timeline(db, case, user, "Note added", payload.note)
    log_action(db, user, "update", "emergency", case.id,
              case.student.full_name if case.student else None, "Note added")
    db.refresh(case)
    return _emergency_out(case)

@app.get("/api/osas/accounts", response_model=List[schemas.OsasAccountOut])
def list_osas_accounts(db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    return db.query(models.User).filter(models.User.role == "osas_admin").all()

@app.put("/api/osas/accounts/{aid}", response_model=schemas.OsasAccountOut)
def update_osas_account(aid: int, payload: schemas.OsasAccountUpdate,
                        db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    a = db.query(models.User).filter(models.User.id == aid, models.User.role == "osas_admin").first()
    if not a: raise HTTPException(404, "Not found")
    for f, v in payload.dict(exclude_unset=True).items(): setattr(a, f, v)
    db.commit(); db.refresh(a)
    log_action(db, user, "update", "osas_account", aid, a.full_name, "Account updated")
    return a

@app.delete("/api/osas/accounts/{aid}")
def delete_osas_account(aid: int, db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    if aid == user.id: raise HTTPException(400, "Cannot delete your own account.")
    a = db.query(models.User).filter(models.User.id == aid, models.User.role == "osas_admin").first()
    if not a: raise HTTPException(404, "Not found")
    log_action(db, user, "delete", "osas_account", aid, a.full_name, "OSAS account deleted")
    db.delete(a); db.commit()
    return {"message": "Deleted"}

@app.get("/api/osas/students", response_model=List[schemas.StudentAccountOut])
def list_students(
    search: Optional[str] = None,
    gender: Optional[str] = None,
    is_archived: Optional[bool] = None,
    course_section: Optional[str] = None,
    db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin),
):
    lifecycle_sweep(db)
    students = db.query(models.User).filter(models.User.role == "student").all()
    results = []
    for s in students:
        if search and search.lower() not in s.full_name.lower() and search.lower() not in s.email.lower(): continue
        if gender and (s.gender or "").lower() != gender.lower(): continue
        if is_archived is not None and (s.archived_at is not None) != is_archived: continue
        if course_section and (s.course_section or "").lower() != course_section.lower(): continue
        results.append(schemas.StudentAccountOut(
            id=s.id, full_name=s.full_name, email=s.email,
            course_section=s.course_section, gender=s.gender, created_at=s.created_at,
            is_archived=s.archived_at is not None, archived_at=s.archived_at,
        ))
    return results

@app.patch("/api/osas/students/{sid}/archive")
def archive_student(sid: int, db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    s = db.query(models.User).filter(models.User.id == sid, models.User.role == "student").first()
    if not s: raise HTTPException(404, "Not found")
    s.archived_at = datetime.utcnow(); db.commit()
    log_action(db, user, "archive", "student", sid, s.full_name, "Student archived")
    return {"message": "Archived"}

@app.patch("/api/osas/students/{sid}/unarchive")
def unarchive_student(sid: int, db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    s = db.query(models.User).filter(models.User.id == sid, models.User.role == "student").first()
    if not s: raise HTTPException(404, "Not found")
    s.archived_at = None; db.commit()
    log_action(db, user, "unarchive", "student", sid, s.full_name, "Student unarchived")
    return {"message": "Unarchived"}

@app.post("/api/osas/announcements")
def send_announcement(payload: schemas.AnnouncementCreate, db: Session = Depends(get_db),
                      user: models.User = Depends(require_osas_admin)):
    if payload.audience not in ("all", "flagged"):
        raise HTTPException(400, "Invalid audience")

    q = db.query(models.User).filter(models.User.role == "student")
    if payload.audience == "flagged":
        flagged_ids = [f.student_id for f in
                       db.query(models.StudentFlag).filter(models.StudentFlag.is_flagged == True).all()]
        q = q.filter(models.User.id.in_(flagged_ids))
    students = q.all()

    sent = 0
    for s in students:
        if email_utils.notify_announcement(s.email, s.full_name, payload.subject, payload.message):
            sent += 1

    notify_all(db, [s.id for s in students], "announcement", payload.subject, payload.message)

    log_action(db, user, "create", "announcement", None, payload.subject,
               f"Sent to {sent}/{len(students)} student(s) ({payload.audience}): {payload.message[:200]}")
    return {"message": "Announcement sent.", "recipients": len(students), "sent": sent}

@app.get("/api/osas/announcements/history")
def announcement_history(db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    logs = (db.query(models.AuditLog)
            .filter(models.AuditLog.resource_type == "announcement")
            .order_by(models.AuditLog.created_at.desc())
            .limit(50).all())
    return [{
        "id": l.id, "subject": l.resource_label, "detail": l.detail,
        "sent_by": l.actor_name, "created_at": l.created_at,
    } for l in logs]

@app.delete("/api/osas/students/{sid}")
def delete_student(sid: int, db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    s = db.query(models.User).filter(models.User.id == sid, models.User.role == "student").first()
    if not s: raise HTTPException(404, "Not found")
    log_action(db, user, "delete", "student", sid, s.full_name, "Student account deleted")
    db.delete(s); db.commit()
    return {"message": "Deleted"}

@app.get("/api/osas/audit-logs", response_model=List[schemas.AuditLogOut])
def get_audit_logs(
    limit: int = 50, action: Optional[str] = None, resource_type: Optional[str] = None,
    db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin),
):
    q = db.query(models.AuditLog).order_by(models.AuditLog.created_at.desc())
    if action: q = q.filter(models.AuditLog.action == action)
    if resource_type: q = q.filter(models.AuditLog.resource_type == resource_type)
    return q.limit(limit).all()


# ─── TALLY / EXPORT ───────────────────────────────────────────────────────────
def _compute_tally(db, group_by_list, month_label):
    valid = {"barangay","boarding_house","gender","department","monthly_status"}
    def compute_section(g):
        groups: dict = {}
        if g in ("gender","department"):
            for s in db.query(models.User).filter(models.User.role=="student").all():
                label = (s.gender or "Not specified").replace("_"," ").title() if g=="gender" else (s.course_section or "Not specified")
                groups.setdefault(label,[]).append(s.full_name)
        else:
            q = db.query(models.StatusUpdate)
            if month_label: q = q.filter(models.StatusUpdate.month_label==month_label)
            for u in q.all():
                nm = u.student.full_name if u.student else "Unknown"
                if g=="monthly_status":
                    label = {"same":"Same boarding house","transferred":"Transferred","moved_home":"Moved home"}.get(u.status_type,u.status_type)
                    groups.setdefault(label,[]).append(nm); continue
                if u.status_type=="transferred": hn,bar = u.new_boarding_house_name or "Unknown",u.new_barangay or "Unknown"
                elif u.boarding_house: hn,bar = u.boarding_house.name,u.boarding_house.barangay
                else: continue
                label = bar if g=="barangay" else hn
                groups.setdefault(label,[]).append(nm)
        rows = [schemas.TallyReportRow(group_label=k,count=len(v),student_names=sorted(v)) for k,v in sorted(groups.items())]
        return schemas.TallyReportSection(group_by=g,rows=rows,total=sum(len(v) for v in groups.values()))
    sections = [compute_section(g) for g in group_by_list if g in valid]
    return sections

@app.get("/api/osas/reports/tally", response_model=schemas.TallyReportOut)
def tally_report(group_by: str, month_label: Optional[str] = None,
                 db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    requested = [g.strip() for g in group_by.split(",") if g.strip()]
    sections = _compute_tally(db, requested, month_label)
    return schemas.TallyReportOut(group_by=",".join(requested), month_label=month_label,
                                  sections=sections, rows=sections[0].rows if sections else [],
                                  total=sections[0].total if len(sections)==1 else sum(s.total for s in sections))

@app.get("/api/osas/reports/export/csv")
def export_csv(group_by: str, month_label: Optional[str] = None,
               db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    import csv, io
    requested = [g.strip() for g in group_by.split(",") if g.strip()]
    sections  = _compute_tally(db, requested, month_label)
    buf = io.StringIO()
    w = csv.writer(buf)
    for sec in sections:
        w.writerow([f"Group by: {sec.group_by}",f"Month: {month_label or 'All'}"])
        w.writerow(["Group", "Count", "Students"])
        for row in sec.rows:
            w.writerow([row.group_label, row.count, "; ".join(row.student_names)])
        w.writerow(["Total", sec.total])
        w.writerow([])
    log_action(db, user, "export", "report", None, group_by, f"CSV export — {month_label or 'all months'}")
    return StreamingResponse(io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition":f"attachment; filename=geotrack_report.csv"})

@app.get("/api/osas/reports/export/excel")
def export_excel(group_by: str, month_label: Optional[str] = None,
                 db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    requested = [g.strip() for g in group_by.split(",") if g.strip()]
    sections  = _compute_tally(db, requested, month_label)
    wb = Workbook()
    for i, sec in enumerate(sections):
        ws = wb.create_sheet(title=sec.group_by[:31]) if i > 0 else wb.active
        ws.title = sec.group_by[:31]
        ws.append([f"GeoTrack Tally Report — {sec.group_by}", f"Month: {month_label or 'All months'}"])
        ws.append(["Group", "Count", "Students"])
        for cell in ws[2]: cell.font = Font(bold=True)
        for row in sec.rows:
            ws.append([row.group_label, row.count, "; ".join(row.student_names)])
        ws.append(["Total", sec.total])
        for col in range(1, 4):
            ws.column_dimensions[get_column_letter(col)].width = [30, 10, 60][col-1]
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    log_action(db, user, "export", "report", None, group_by, f"Excel export — {month_label or 'all months'}")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":"attachment; filename=geotrack_report.xlsx"})

@app.get("/api/osas/reports/export/pdf")
def export_pdf(group_by: str, month_label: Optional[str] = None,
               db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
    from reportlab.lib import colors
    requested = [g.strip() for g in group_by.split(",") if g.strip()]
    sections  = _compute_tally(db, requested, month_label)
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = [Paragraph("GeoTrack Tally Report", styles["Title"]),
             Paragraph(f"Month: {month_label or 'All months'} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC", styles["Normal"]),
             Spacer(1, 16)]
    for sec in sections:
        story.append(Paragraph(f"Group by: {sec.group_by}", styles["Heading2"]))
        data = [["Group", "Count", "Students"]]
        for row in sec.rows:
            data.append([row.group_label, str(row.count), "\n".join(row.student_names)])
        data.append(["Total", str(sec.total), ""])
        t = Table(data, colWidths=[130, 50, 320])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#203f36")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.white),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f6f4ee")]),
            ("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"),
            ("GRID",       (0,0), (-1,-1), 0.5, colors.HexColor("#d9d3c4")),
            ("FONTSIZE",   (0,0), (-1,-1), 9),
            ("VALIGN",     (0,0), (-1,-1), "TOP"),
        ]))
        story += [t, Spacer(1, 16)]
    doc.build(story)
    buf.seek(0)
    log_action(db, user, "export", "report", None, group_by, f"PDF export — {month_label or 'all months'}")
    return StreamingResponse(buf, media_type="application/pdf",
        headers={"Content-Disposition":"attachment; filename=geotrack_report.pdf"})

@app.get(
    "/api/student/compliance",
    response_model=List[schemas.StudentComplianceOut]
)
def student_compliance(
    db: Session = Depends(get_db),
    user: models.User = Depends(require_student)
):
    return (
        db.query(models.StudentCompliance)
        .filter(models.StudentCompliance.student_id == user.id)
        .order_by(
            models.StudentCompliance.year.desc(),
            models.StudentCompliance.id.desc()
        )
        .all()
    )
@app.get(
    "/api/student/compliance/status",
    response_model=schemas.StudentFlagOut
)
def compliance_status(
    db: Session = Depends(get_db),
    user: models.User = Depends(require_student)
):
    flag = (
        db.query(models.StudentFlag)
        .filter(models.StudentFlag.student_id == user.id)
        .first()
    )

    if not flag:
        flag = models.StudentFlag(
            student_id=user.id
        )
        db.add(flag)
        db.commit()
        db.refresh(flag)

    return flag


@app.post("/api/osas/compliance/send-reminders")
def send_compliance_reminders(
    db: Session = Depends(get_db),
    user: models.User = Depends(require_osas_admin)
):
    now = datetime.utcnow()
    current_month = now.strftime("%B")
    current_year = now.year
    month_label = f"{current_month} {current_year}"
    is_final = now.day >= COMPLIANCE_DEADLINE_DAY - 2

    students = db.query(models.User).filter(
        models.User.role == "student"
    ).all()

    reminders = []

    for student in students:

        compliance = (
            db.query(models.StudentCompliance)
            .filter(
                models.StudentCompliance.student_id == student.id,
                models.StudentCompliance.month == current_month,
                models.StudentCompliance.year == current_year,
            )
            .first()
        )

        if not compliance:
            compliance = models.StudentCompliance(
                student_id=student.id,
                month=current_month,
                year=current_year,
                submission_status="Pending",
                deadline=datetime(current_year, now.month, COMPLIANCE_DEADLINE_DAY),
            )
            db.add(compliance)
            db.commit()

        if compliance.submission_status != "Pending":
            continue

        if is_final:
            sent = email_utils.notify_final_reminder(student.email, student.full_name, month_label)
            notify(db, student.id, "final_reminder", "Final reminder: submission due soon",
                   f"Your {month_label} status update is still pending and the deadline is near.")
        else:
            sent = email_utils.notify_monthly_reminder(
                student.email, student.full_name, month_label, COMPLIANCE_DEADLINE_DAY)
            notify(db, student.id, "monthly_reminder", "Monthly submission reminder",
                   f"Please submit your {month_label} status update before day {COMPLIANCE_DEADLINE_DAY}.")

        reminders.append({
            "student": student.full_name,
            "email": student.email,
            "type": "final" if is_final else "monthly",
            "sent": sent,
        })

    db.commit()
    log_action(db, user, "create", "compliance", None, None,
               f"Sent {sum(1 for r in reminders if r['sent'])} {'final' if is_final else 'monthly'} reminder(s)")

    return {
        "message": "Reminder process completed.",
        "total_reminders": len(reminders),
        "students": reminders
    }

@app.post("/api/osas/compliance/check-missed")
def check_missed_submissions(
    db: Session = Depends(get_db),
    user: models.User = Depends(require_osas_admin)
):
    now = datetime.utcnow()

    compliances = db.query(models.StudentCompliance).filter(
        models.StudentCompliance.submission_status == "Pending"
    ).all()

    missed_count = 0

    for compliance in compliances:

        if compliance.deadline and now > compliance.deadline:

            compliance.submission_status = "Missed"
            compliance.remarks = "Submission deadline missed"

            flag = db.query(models.StudentFlag).filter(
                models.StudentFlag.student_id == compliance.student_id
            ).first()

            if not flag:

                flag = models.StudentFlag(
                    student_id=compliance.student_id,
                    missed_count=1,
                    is_flagged=False
                )

                db.add(flag)

            else:

                 flag.missed_count += 1

            was_flagged = flag.is_flagged
            if flag.missed_count >= 3:
                flag.is_flagged = True
                flag.compliance_status = "Flagged"
                flag.reason = "Three consecutive missed submissions"
                if not was_flagged:
                    db.commit()
                    notify(db, compliance.student_id, "flagging", "You've been flagged for compliance",
                           f"You have {flag.missed_count} missed submission(s) and have been flagged by OSAS.")

            missed_count += 1

    db.commit()

    return {
        "message": "Missed submission checking completed.",
        "updated_records": missed_count
    }

@app.get("/api/osas/compliance/history")
def compliance_history(
    student_id: Optional[int] = None,
    month: Optional[str] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(require_osas_admin)
):
    query = (
        db.query(models.StudentCompliance)
        .join(models.User)
        .filter(models.User.role == "student")
    )

    if student_id:
        query = query.filter(models.StudentCompliance.student_id == student_id)

    if month:
        query = query.filter(models.StudentCompliance.month == month)

    if year:
        query = query.filter(models.StudentCompliance.year == year)

    records = query.order_by(
        models.StudentCompliance.year.desc(),
        models.StudentCompliance.id.desc()
    ).all()

    results = []

    for record in records:

        flag = (
            db.query(models.StudentFlag)
            .filter(models.StudentFlag.student_id == record.student_id)
            .first()
        )

        results.append({
            "student_id": record.student_id,
            "student_name": record.student.full_name,
            "email": record.student.email,
            "course_section": record.student.course_section,
            "month": record.month,
            "year": record.year,
            "submission_status": record.submission_status,
            "submitted_at": record.submitted_at,
            "deadline": record.deadline,
            "remarks": record.remarks,
            "is_flagged": flag.is_flagged if flag else False,
            "missed_submissions": flag.missed_count if flag else 0
        })

    return results

@app.post("/api/osas/compliance/update-flags")
def update_student_flags(
    db: Session = Depends(get_db),
    user: models.User = Depends(require_osas_admin)
):
    flags_updated = 0

    students = db.query(models.User).filter(
        models.User.role == "student"
    ).all()

    for student in students:

        missed = db.query(models.StudentCompliance).filter(
            models.StudentCompliance.student_id == student.id,
            models.StudentCompliance.submission_status == "Missed"
        ).count()

        flag = db.query(models.StudentFlag).filter(
            models.StudentFlag.student_id == student.id
        ).first()

        if not flag:

            flag = models.StudentFlag(
                student_id=student.id,
                missed_count=missed,
                is_flagged=False,
                reason=None
            )

            db.add(flag)

        else:

            flag.missed_count = missed

        was_flagged = flag.is_flagged
        if missed >= 3:
            flag.is_flagged = True
            flag.compliance_status = "Flagged"
            flag.reason = "Student missed three or more monthly submissions."
            if not was_flagged:
                db.commit()
                notify(db, student.id, "flagging", "You've been flagged for compliance",
                       f"You have {missed} missed submission(s) and have been flagged by OSAS.")
        else:
            flag.is_flagged = False
            flag.compliance_status = "Good"
            flag.reason = None

        flags_updated += 1

    db.commit()

    return {
        "message": "Student flags updated successfully.",
        "students_checked": flags_updated
    }

@app.get("/api/osas/compliance/report")
def compliance_report(
    month: Optional[str] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    user: models.User = Depends(require_osas_admin)
):
    now = datetime.utcnow()

    month = month or now.strftime("%B")
    year = year or now.year

    compliances = db.query(models.StudentCompliance).filter(
        models.StudentCompliance.month == month,
        models.StudentCompliance.year == year
    ).all()

    results = []

    for compliance in compliances:

        student = compliance.student

        flag = db.query(models.StudentFlag).filter(
            models.StudentFlag.student_id == student.id
        ).first()

        results.append({
            "student_name": student.full_name,
            "email": student.email,
            "course_section": student.course_section,
            "month": compliance.month,
            "year": compliance.year,
            "status": compliance.submission_status,
            "submitted_at": compliance.submitted_at,
            "deadline": compliance.deadline,
            "remarks": compliance.remarks,
            "flagged": flag.is_flagged if flag else False
        })

    return results

# ─── Notifications ────────────────────────────────────────────────────────────
@app.get("/api/student/notifications", response_model=List[schemas.NotificationOut])
def student_notifications(db: Session = Depends(get_db), user: models.User = Depends(require_student)):
    return (db.query(models.Notification)
            .filter(models.Notification.user_id == user.id)
            .order_by(models.Notification.created_at.desc()).all())

@app.patch("/api/student/notifications/{nid}/read")
def student_mark_notification_read(nid: int, db: Session = Depends(get_db),
                                   user: models.User = Depends(require_student)):
    n = db.query(models.Notification).get(nid)
    if not n or n.user_id != user.id: raise HTTPException(404, "Not found")
    n.is_read = True; db.commit()
    return {"message": "Marked as read"}

@app.patch("/api/student/notifications/read-all")
def student_mark_all_notifications_read(db: Session = Depends(get_db),
                                        user: models.User = Depends(require_student)):
    db.query(models.Notification).filter(
        models.Notification.user_id == user.id, models.Notification.is_read == False
    ).update({"is_read": True})
    db.commit()
    return {"message": "All marked as read"}

@app.get("/api/osas/notifications", response_model=List[schemas.NotificationOut])
def osas_notifications(db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    return (db.query(models.Notification)
            .filter(models.Notification.user_id == user.id)
            .order_by(models.Notification.created_at.desc()).all())

@app.patch("/api/osas/notifications/{nid}/read")
def osas_mark_notification_read(nid: int, db: Session = Depends(get_db),
                                user: models.User = Depends(require_osas_admin)):
    n = db.query(models.Notification).get(nid)
    if not n or n.user_id != user.id: raise HTTPException(404, "Not found")
    n.is_read = True; db.commit()
    return {"message": "Marked as read"}

@app.patch("/api/osas/notifications/read-all")
def osas_mark_all_notifications_read(db: Session = Depends(get_db),
                                     user: models.User = Depends(require_osas_admin)):
    db.query(models.Notification).filter(
        models.Notification.user_id == user.id, models.Notification.is_read == False
    ).update({"is_read": True})
    db.commit()
    return {"message": "All marked as read"}


# ─── Compliance: combined automation sweep ────────────────────────────────────
def run_compliance_automation(db: Session, actor: Optional[models.User] = None):
    """Runs the full daily sweep: ensure this month's records exist, detect
    missed submissions, recompute flags, and send reminder emails/notifications.
    Used by both the manual 'Run automation now' button and the scheduler."""
    check_monthly_compliance(db)

    now = datetime.utcnow()
    compliances = db.query(models.StudentCompliance).filter(
        models.StudentCompliance.submission_status == "Pending"
    ).all()
    missed_count = 0
    for compliance in compliances:
        if compliance.deadline and now > compliance.deadline:
            compliance.submission_status = "Missed"
            compliance.remarks = "Submission deadline missed"
            flag = db.query(models.StudentFlag).filter(
                models.StudentFlag.student_id == compliance.student_id).first()
            if not flag:
                flag = models.StudentFlag(student_id=compliance.student_id, missed_count=1, is_flagged=False)
                db.add(flag)
            else:
                flag.missed_count += 1
            was_flagged = flag.is_flagged
            if flag.missed_count >= 3:
                flag.is_flagged = True
                flag.compliance_status = "Flagged"
                flag.reason = "Three consecutive missed submissions"
                if not was_flagged:
                    db.commit()
                    notify(db, compliance.student_id, "flagging", "You've been flagged for compliance",
                           f"You have {flag.missed_count} missed submission(s) and have been flagged by OSAS.")
            missed_count += 1
    db.commit()

    students = db.query(models.User).filter(models.User.role == "student").all()
    flags_updated = 0
    for student in students:
        missed = db.query(models.StudentCompliance).filter(
            models.StudentCompliance.student_id == student.id,
            models.StudentCompliance.submission_status == "Missed").count()
        flag = db.query(models.StudentFlag).filter(models.StudentFlag.student_id == student.id).first()
        if not flag:
            flag = models.StudentFlag(student_id=student.id, missed_count=missed, is_flagged=False)
            db.add(flag)
        else:
            flag.missed_count = missed
        was_flagged = flag.is_flagged
        if missed >= 3:
            flag.is_flagged = True
            flag.compliance_status = "Flagged"
            flag.reason = "Student missed three or more monthly submissions."
            if not was_flagged:
                db.commit()
                notify(db, student.id, "flagging", "You've been flagged for compliance",
                       f"You have {missed} missed submission(s) and have been flagged by OSAS.")
        else:
            flag.is_flagged = False
            flag.compliance_status = "Good"
            flag.reason = None
        flags_updated += 1
    db.commit()

    current_month = now.strftime("%B")
    current_year = now.year
    month_label = f"{current_month} {current_year}"
    is_final = now.day >= COMPLIANCE_DEADLINE_DAY - 2
    reminders_sent = 0
    for student in students:
        compliance = db.query(models.StudentCompliance).filter(
            models.StudentCompliance.student_id == student.id,
            models.StudentCompliance.month == current_month,
            models.StudentCompliance.year == current_year).first()
        if not compliance or compliance.submission_status != "Pending":
            continue
        if is_final:
            sent = email_utils.notify_final_reminder(student.email, student.full_name, month_label)
            notify(db, student.id, "final_reminder", "Final reminder: submission due soon",
                   f"Your {month_label} status update is still pending and the deadline is near.")
        else:
            sent = email_utils.notify_monthly_reminder(
                student.email, student.full_name, month_label, COMPLIANCE_DEADLINE_DAY)
            notify(db, student.id, "monthly_reminder", "Monthly submission reminder",
                   f"Please submit your {month_label} status update before day {COMPLIANCE_DEADLINE_DAY}.")
        if sent: reminders_sent += 1
    db.commit()

    log_action(db, actor, "create", "compliance", None, None,
               f"Automation sweep: {missed_count} newly missed, {flags_updated} flags recomputed, "
               f"{reminders_sent} reminder(s) sent")

    return {
        "message": "Automation sweep completed.",
        "newly_missed": missed_count,
        "flags_recomputed": flags_updated,
        "reminders_sent": reminders_sent,
    }


@app.post("/api/osas/compliance/run-automation")
def run_automation_now(db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    return run_compliance_automation(db, actor=user)


# ─── Risk assessment ───────────────────────────────────────────────────────────
RISK_WEIGHTS = {"missed": 12, "emergency": 15, "concern": 10}

def _risk_level(score: int) -> str:
    if score >= 70: return "Critical"
    if score >= 45: return "High"
    if score >= 20: return "Medium"
    return "Low"

@app.get("/api/osas/risk-assessment", response_model=List[schemas.RiskAssessmentRow])
def risk_assessment(db: Session = Depends(get_db), user: models.User = Depends(require_osas_admin)):
    students = db.query(models.User).filter(models.User.role == "student").all()
    rows = []
    for s in students:
        flag = db.query(models.StudentFlag).filter(models.StudentFlag.student_id == s.id).first()
        missed = flag.missed_count if flag else 0
        emergencies = db.query(models.EmergencyCase).filter(models.EmergencyCase.student_id == s.id).count()
        open_concerns = db.query(models.Concern).filter(
            models.Concern.student_id == s.id, models.Concern.status != "resolved").count()

        raw_score = (missed * RISK_WEIGHTS["missed"] + emergencies * RISK_WEIGHTS["emergency"]
                     + open_concerns * RISK_WEIGHTS["concern"])
        risk_score = min(100, raw_score)
        compliance_score = max(0, 100 - missed * 20)

        rows.append(schemas.RiskAssessmentRow(
            student_id=s.id, student_name=s.full_name, email=s.email,
            course_section=s.course_section, missed_submissions=missed,
            emergency_count=emergencies, open_concern_count=open_concerns,
            compliance_score=compliance_score, risk_score=risk_score,
            risk_level=_risk_level(risk_score),
        ))

    rows.sort(key=lambda r: r.risk_score, reverse=True)
    return rows


@app.get("/api/")
def root():
    return {"message": "GeoTrack API v3 — /docs for explorer."}